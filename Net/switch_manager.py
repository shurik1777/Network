# Импорт необходимых библиотек
import openpyxl  # Для работы с Excel-файлами (установка: pip install openpyxl)
import re  # Для работы с регулярными выражениями
from netmiko import ConnectHandler  # Для установки SSH-соединений с устройствами (установка: pip install netmiko)
from pathlib import Path  # Для работы с путями к файлам и директориям
from openpyxl.styles import Alignment, Font  # Для настройки стилей ячеек в Excel


class NetworkDevice:
    def __init__(self, device_info):
        self.device_info = device_info
        self.connection = None

    def establish_connection(self):
        self.connection = ConnectHandler(**self.device_info)

    def disconnect(self):
        if self.connection:
            self.connection.disconnect()

    def send_command(self, command):
        return self.connection.send_command(command)


# Класс для управления коммутаторами
class SwitchManager:
    def __init__(self, device_info):
        self.device_info = device_info
        self.network_device = NetworkDevice(device_info)

    def establish_connection(self):
        """Установка соединения с устройством."""
        connection = ConnectHandler(**self.device_info)
        return connection

    def get_switch_ports(self, connection):
        """Получение информации о портах коммутатора с устройства."""
        switch_ports = connection.send_command('show ports')
        connection.disconnect()
        return switch_ports

    def get_switch_info(self, connection):
        """Получение общей информации о коммутаторе с устройства."""
        switch_info = connection.send_command('show switch')
        connection.disconnect()
        return switch_info

    def get_vlan_info(self, connection):
        """Получение общей информации о коммутаторе с устройства."""
        vlan_info = connection.send_command('show vlan')
        connection.disconnect()
        return vlan_info

    def parse_ports(self, switch_ports):
        """Анализ сырых данных о портах коммутатора и преобразование их в список отформатированных строк."""
        lines = switch_ports.strip().split('\n')
        parsed_data = []
        for line in lines[4:-1]:
            if line.strip():
                port_data = re.split(r'\s{2,}', line.strip())
                filtered_port_data = [item for item in port_data if item != "Auto"]
                parsed_data.append(filtered_port_data)
        return parsed_data

    def save_to_excel(self, info_output, ports_output, vlan_data_list):
        """Создание книги Excel и сохранение отформатированных данных в листах."""
        # Создаем книгу Excel
        workbook = openpyxl.Workbook()

        # Лист для информации о коммутаторе
        info_sheet = workbook.active
        info_sheet.title = 'Информация о коммутаторе'

        # Заполняем информацию о коммутаторе
        info_lines = info_output.strip().split('\n')
        for row, line in enumerate(info_lines, start=1):
            # Пропускаем первую строку с "Show switch"
            if row == 1:
                continue

            line_parts = line.split(':', 1)
            if len(line_parts) >= 2:
                key, value = line_parts
                info_sheet.cell(row=row - 1, column=1, value=key.strip())
                info_sheet.cell(row=row - 1, column=2, value=value.strip())

        # Лист для информации о портах
        ports_sheet = workbook.create_sheet(title='Порты')

        # Форматирование для заголовков
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_font = Font(name='Arial', size=12, bold=True)
        header_body = Font(name='Arial', size=12)

        header_row = ["Порт", "Состояние/MDI", "Настройки Скорость/Duplex/FlowCtrl",
                      "Соединение Скорость/Duplex/FlowCtrl", "Обучение адресов"]
        ports_sheet.append(header_row)
        for col_num, cell in enumerate(ports_sheet[1], start=1):
            column_letter = openpyxl.utils.get_column_letter(col_num)
            ports_sheet.column_dimensions[column_letter].width = 25
            cell.alignment = header_alignment
            cell.font = header_font

        # Заполняем информацию о портах
        parsed_ports = self.parse_ports(ports_output)
        for parsed_row in parsed_ports:
            ports_sheet.append(parsed_row)

        # Лист для информации о VLAN
        vlan_sheet = workbook.create_sheet(title='VLAN')

        # Заголовки для VLAN
        vlan_header_row = ["VLAN ID", "Имя", "Тегированные порты", "Нетегированные порты"]
        vlan_sheet.append(vlan_header_row)
        for col_num, cell in enumerate(vlan_sheet[1], start=1):
            column_letter = openpyxl.utils.get_column_letter(col_num)
            vlan_sheet.column_dimensions[column_letter].width = 25
            cell.alignment = header_alignment
            cell.font = header_font

        # Заполняем информацию о VLAN
        for vlan_data in vlan_data_list:
            parsed_row = [vlan_data.get('Vlan ID', ''), vlan_data.get('Name', ''),
                          vlan_data.get('Tagged Ports', ''), vlan_data.get('Untagged Ports', '')]
            vlan_sheet.append(parsed_row)

        # Форматируем строки в листе "VLAN"
        for row in vlan_sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = header_alignment
                cell.font = header_body

        # Форматируем заголовок "Порты"
        header_row = ["Порт", "Состояние/MDI", "Настройки Скорость/Duplex/FlowCtrl",
                      "Соединение Скорость/Duplex/FlowCtrl",
                      "Обучение адресов"]

        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        font = Font(name='Arial', size=12)
        max_column_widths = [25] * len(header_row)

        # Форматируем строки в листе "Информация о коммутаторе"
        info_sheet = workbook['Информация о коммутаторе']
        for col_num, width in enumerate(max_column_widths, start=1):
            column_letter = openpyxl.utils.get_column_letter(col_num)
            info_sheet.column_dimensions[column_letter].width = width + 2
            for row in info_sheet.iter_rows(min_row=1, max_row=info_sheet.max_row, min_col=col_num, max_col=col_num):
                for cell in row:
                    new_alignment = Alignment(horizontal=center_alignment.horizontal,
                                              vertical=center_alignment.vertical,
                                              wrap_text=True)
                    cell.alignment = new_alignment
                    cell.font = font

        # Форматируем строки в листе "Порты"
        ports_sheet = workbook['Порты']
        for row in ports_sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = center_alignment
                cell.font = font

        return workbook

    def parse_vlan(self, vlan_info):
        """Анализ сырых данных о VLAN и преобразование их в список отформатированных строк."""
        lines = vlan_info.strip().split('\n')
        parsed_data = []
        vlan_data = {}

        for line in lines:
            if line.startswith('VID'):
                if vlan_data:
                    parsed_data.append(vlan_data)
                vlan_data = {}

                vid_match = re.search(r'VID\s+:\s+(\d+)', line)
                if vid_match:
                    vlan_data['Vlan ID'] = vid_match.group(1)

                name_match = re.search(r'VLAN Name\s+:\s+(.+)', line)
                if name_match:
                    vlan_data['Name'] = name_match.group(1)

            elif line.startswith('Current Tagged Ports'):
                tagged_match = re.search(r'Current Tagged Ports\s+:\s+(.+)', line)
                if tagged_match:
                    vlan_data['Tagged Ports'] = tagged_match.group(1)

            elif line.startswith('Current Untagged Ports'):
                untagged_match = re.search(r'Current Untagged Ports\s+:\s+(.+)', line)
                if untagged_match:
                    vlan_data['Untagged Ports'] = untagged_match.group(1)

        # Добавляем последний vlan_data, который не был добавлен в цикле
        if vlan_data:
            parsed_data.append(vlan_data)

        return parsed_data


def run_switch_manager(ip, username, password):
    devices = [
        {'ip': ip, 'username': username, 'password': password, 'device_type': 'dlink_ds_ssh'},
        # Добавьте остальные коммутаторы в список
    ]
    for device_info in devices:
        switch_manager = SwitchManager(device_info)

        network_device = switch_manager.network_device

        network_device.establish_connection()

        switch_info_output = network_device.send_command('show switch')
        switch_ports_output = network_device.send_command('show ports')
        vlan_info_output = network_device.send_command('show vlan')

        network_device.disconnect()

        # Парсинг информации о VLAN
        parsed_vlans = switch_manager.parse_vlan(vlan_info_output)

        # Сохранение в Excel-файл
        workbook = switch_manager.save_to_excel(switch_info_output, switch_ports_output, parsed_vlans)

        file_number = 1
        file_name = f'switch_info_{device_info["ip"]}_{file_number}.xlsx'
        while Path(file_name).is_file():
            file_number += 1
            file_name = f'switch_info_{device_info["ip"]}_{file_number}.xlsx'

        workbook.save(file_name)
